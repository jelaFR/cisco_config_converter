def customer_logo(sleep_time):
    """Display Customer logo as welcome message"""
    from time import sleep
    import os

    logo = [""]
    # Displaying logo
    os.system("cls")
    for line in logo:
        print(line)
        sleep(sleep_time)


def exception_logo():
    """Display error message in case of crash"""

    print("""
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,:do:,,;oOo;,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,cxdc,,ck0l,,,,,,,,,,,,:dK0xd0Xkc,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,:d0KkkK0o:,,,,,,,,,,,,,,:kNWNk:,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,cOWWNx:,,,,,,,,,,,,,,:dK0xOK0o:,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,;lOXkokK0d;,,,,,,,,,,,,o0x:,;cxxc,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,;dkl,,,:oo,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,;::ccclcc::;,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,:ldxO0KXNNNNNNXK0Okdoc;,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,;cdOKNXK0kxxddddddddxkO0KK0ko:;,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,:oOKKOxlc;,,,,,,,,,,,,,,,;:coxOkc,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,:kKOo:,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    ,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,
    """)


def search_files_in_dir(dir_path):
    """List all file in directory"""
    import os
    files = list()
    
    # r = root, d = directory, f = files
    for r, d, f in os.walk(dir_path):
        for file in f:
            files.append(os.path.join(r, file))
    return files


def process_file(cfg_params, out_path, matrix):
    """
        Launch all tasks required to convert config files
        :return file_association: (name_of_old_config, name_of_new_config, device_type) or False
    """
    import os

    cfg_file, dev_type = cfg_params
    
    # Convert file according to its type
    new_file = convert_cfg_file(cfg_file, dev_type, out_path, matrix)
    # Output var
    file_association = (cfg_file, new_file, dev_type)

    return file_association


def check_cfg_file(cfg_file):
    """ 
        Read Cisco configuration file and check if this is valid
    """
    from ciscoconfparse import CiscoConfParse
    import re
    # Initiate vars
    file_valid = False

    # Check if file is valid to avoid unnecessary processing
    try:
        config_parse = CiscoConfParse(cfg_file)
        # Check if hostname appears in file
        hostname = config_parse.find_objects(r"^hostname ")
        if len(hostname) != 0:
            hostname = hostname[0].text.replace("hostname ","")
            file_valid = True

    except UnicodeDecodeError: # This occurs if file is not a text file
        pass

    return file_valid


def check_dev_model(cfg_file):
    """
        Try to determine model based on licence udi or number of ifaces
    """
    from ciscoconfparse import CiscoConfParse
    import re

    config_parse = CiscoConfParse(cfg_file)


    # Check license udi PID to determine model
    dev_model = config_parse.find_objects(r"^license udi pid")

    if len(dev_model) != 0: # As there is a licence provided in config file, it's easy to figure the model
        dev_model = dev_model[0].text
        dev_model = dev_model.split(" ")[3].replace("/K9","").replace("ISR","").replace("CISCO","").replace("DC","").replace("-","")
    else: # Looking for clue in order to determine switch model
        switch_fa_ifaces = config_parse.find_objects(r"^interface FastEthernet0/")
        switch_gi_ifaces = config_parse.find_objects(r"^interface GigabitEthernet0/")
        # Check if this is a switch card
        if len(switch_fa_ifaces) == 23 and len(switch_gi_ifaces) == 3:
            dev_model = "SM-ES3-24-P"
        
    # Do not convert other devices
    if not(dev_model == "2951" or dev_model == "2911" or dev_model == "1921" or dev_model == "SM-ES3-24-P"):
        dev_model = "Unknown"

    return dev_model


def check_xlsx_file(xlsx_file):
    """ Check if this is a valid excel file """
    from openpyxl import load_workbook
    from openpyxl import worksheet
    from openpyxl.worksheet.table import Table, TableStyleInfo
    
    # Read excel file
    wb = load_workbook(xlsx_file, read_only=True)
    matrix_sheet = wb.active

    # Check Excel structure
    if "matrix" not in wb.sheetnames: # A sheet named matrix exist
        print("[ERROR]Excel file is not valid, initial sheet name does not exist !")
        is_valid = False
    elif matrix_sheet.max_column != 6: # Sheet contains 6 columns
        print("[ERROR]Excel file is not valid, nb_columns is invalid !")
        is_valid = False
    else:
        is_valid = True

    return is_valid


def convert_cfg_file(config, device_type, out_path, conversion_matrix):
    """Convert cfg file to other cfg file"""

    import os
    import re
    from ciscoconfparse import CiscoConfParse

    # Check if device type exist in conversion matrix
    if device_type in conversion_matrix:
        # Determine new filename
        new_filename = os.path.join(out_path,os.path.basename(config))
        if os.path.isfile(new_filename): # Remove CFG if it exist
            os.remove(new_filename)

        # Parse cisco configuration with Ciscoconfparse
        parse = CiscoConfParse(config)

        # DELETE
        for item in conversion_matrix[device_type]["delete"]:
            if item[1] == None: # Check required fields
                continue
            elif item[0] != None: # Parent cmd is mentionned
                parent_object = parse.find_objects(item[0])
                for parent in parent_object:
                    # Delete child object in parent object
                    parent.delete_children_matching(item[1])
            else: # parent cmd is not mentionned
                cli_objects = parse.find_objects(item[1])
                for cli_object in cli_objects:
                    # Delete object and all child objects if exist
                    cli_object.delete()

        # ADD
        for item in conversion_matrix[device_type]["add"]:
            if item[2] == None: # Check required fields
                continue
            elif item[0] != None: # parent cmd is mentionned
                parent_object = parse.find_objects(item[0])
                parent_object_done = list() # This is to avoid duplicate added entries
                for parent in parent_object:
                    parent_re = re.compile(parent.text)
                    if parent.has_children == True: # Add space to child if they are child
                        if parent.text not in parent_object_done: # Avoid duplicates entries
                            nb_space = len(parent.text) - len(parent.text.lstrip()) + 1
                            parse.insert_after(parent_re,insertstr=" "*nb_space+item[2])
                            parent_object_done.append(parent.text) 
                    else: # Entry is at the root of cfg, no space added
                        parse.insert_after(parent_re,insertstr=item[2])
            else: # parent cmd is not mentionned
                parse.append_line(item[2]) # Write line at the end of the file

        # REPLACE
        for item in conversion_matrix[device_type]["replace"]:
            if item[1] == None or item[2] == None: # Check required fields
                continue
            if item[0] != None: # parent cmd is mentionned
                initial_cmd = re.compile(item[1])
                parse.replace_children(item[0], initial_cmd, item[2])
            else: # parent cmd is not mentionned
                initial_cmd = re.compile(item[1])
                parse.replace_lines(initial_cmd,item[2])
        
        # Write output to out_file
        parse.save_as(new_filename)
    else:
        new_filename = "Skipped (model unknown)"

    return new_filename


def collect_files(args):
    """
        Collect arguments if provided or ask user to provide files
    """
    import argparse
    import os

    # Initiate vars to check what args user gave
    in_filename_provided = False
    in_matrix_provided = False
    out_path_provided = False

    # Case 1 : User provide arguments

    # IN_PATH
    if args.in_path and os.path.isfile(args.in_path): # user provide file to be converted
        in_filename = [args.in_path]
        in_filename_provided = True
        in_filename_type = "file"

    elif args.in_path and os.path.isdir(args.in_path): # user provide directory containing files to be converted
        in_filename = search_files_in_dir(args.in_path)
        in_filename_provided = True
        in_filename_type = "dir"

    # IN_MATRIX
    if args.in_matrix and os.path.isfile(args.in_matrix): # user provide matrix file
        in_matrix = args.in_matrix
        matrix_is_valid = check_xlsx_file(in_matrix)
        if matrix_is_valid == True:
            in_matrix_provided = True

    # OUT_PATH
    if args.out_path and os.path.isdir(args.out_path): # user provide directory where out files will be stored
        out_path = args.out_path
        out_path_provided = True


    # Case 2 : User does not provide argument (or arguments are invalid)

    # Prompt user for input path or file
    while in_filename_provided == False:
        in_filename = input("\nPlease provide path to config files: ")

        if os.path.isfile(in_filename): # user provide file to be converted
            in_filename_provided = True
            in_filename_type = "file"
            break # end of loop

        elif os.path.isdir(in_filename): # user provide directory containing files to be converted
            in_filename = search_files_in_dir(in_filename)
            in_filename_provided = True
            in_filename_type = "dir"
            break # end of loop

        else: # user does not provide anything, looping
            print(f"[ERROR] {in_filename} is neither a file not a directory !")
            in_filename_provided = False

    # Prompt user for input matrix file
    while in_matrix_provided == False:
        in_matrix = input("\nPlease enter the name of the Excel conversion matrix file [*.xlsx]: ")

        if os.path.isfile(in_matrix): # user provide correct matrix file
            matrix_is_valid = check_xlsx_file(in_matrix)
            if matrix_is_valid == True:
                in_matrix_provided = True
                break # end of loop

        else:
            print(f"[ERROR] {in_matrix} is not a file !")
            in_matrix_provided = False

    # Prompt user for output path
    while out_path_provided == False:
        out_path = input("\nPlease type the destination folder path where config files will be written: ")

        if os.path.isdir(out_path): # user provide correct output path
            out_path_provided = True
            break # end of loop

        else:
            print(f"[ERROR] {out_path} is not a directory !")
            out_path_provided = False

            
    return in_filename, in_filename_type, in_matrix, out_path


def process_xlsx_file(xlsx_file):
    """
        Analyse excel file and build a dictionary based on this file
    """
    from openpyxl import load_workbook
    import re

    # Initiate out dict
    matrix = dict()
    matrix["2951"] = dict()
    matrix["2951"]["add"] = list()
    matrix["2951"]["replace"] = list()
    matrix["2951"]["delete"] = list()

    matrix["2911"] = dict()
    matrix["2911"]["add"] = list()
    matrix["2911"]["replace"] = list()
    matrix["2911"]["delete"] = list()

    matrix["1921"] = dict()
    matrix["1921"]["add"] = list()
    matrix["1921"]["replace"] = list()
    matrix["1921"]["delete"] = list()

    matrix["SM-ES3-24-P"] = dict()
    matrix["SM-ES3-24-P"]["add"] = list()
    matrix["SM-ES3-24-P"]["replace"] = list()
    matrix["SM-ES3-24-P"]["delete"] = list()

    # Read excel file
    wb = load_workbook(xlsx_file, read_only=True)
    sheet = wb.active
    for row in range(12, sheet.max_row):
        # Initiate vars
        device = sheet["A"+str(row)].value
        action = sheet["B"+str(row)].value
        parent_cmd = sheet["C"+str(row)].value if sheet["C"+str(row)].value != None else None
        initial_cmd = sheet["D"+str(row)].value if sheet["D"+str(row)].value != None else None
        final_cmd = sheet["E"+str(row)].value if sheet["E"+str(row)].value != None else None

        
        # Do not consider empty lines
        if ((device == None) or (action == None)):
            continue

        # Means that line contains data
        else:
            # Excel structure is valid, processing data
            if device == "All devices": # Means that all devices are concerned
                for device_type in matrix:
                    matrix[device_type][action].append((parent_cmd, initial_cmd, final_cmd))
            elif device == "All routers": # Means that all routers are concerned
                for device_type in ("2951","2911","1921"):
                    matrix[device_type][action].append((parent_cmd, initial_cmd, final_cmd))
            else: # Means that only one device is concerned
                matrix[str(device)][action].append((parent_cmd, initial_cmd, final_cmd))

    return matrix


def display_new_files(new_files):
    """
        Display new files created and properties
    """
    # Display converted files
    if len(new_files) > 0: # There is a converted file
        end_msg="End of file conversion, here is the list of converted file(s):"
        print("\n", "*"*len(end_msg), "\n", end_msg, "\n", "*"*len(end_msg), "\n")
        for config_params in new_files:
            # Device_type : initial filename -> dest_filename
            print(f"{config_params[2]} : {config_params[0]} -> {config_params[1]}")
    else: # No file has been converted
        end_msg="Conversion completed, but no file has been converted !"
        print("\n", "*"*len(end_msg), "\n", end_msg, "\n", "*"*len(end_msg), "\n")


if __name__ == "__main__":
    import argparse
    import os
    import sys
    from time import sleep
    from tqdm import tqdm
    
    # Tell the script what to do with args
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--in_path", help="full path where config file(s) is(are) located")
    parser.add_argument("-o", "--out_path", help="full path where converted files will be located")
    parser.add_argument("-m", "--in_matrix", help="path where XLSX convesion matrix file is located")
    args = parser.parse_args()
    new_files = list()
    valid_files = list()

    header = "\nCisco configuration converter \n\
            Version 1.0\n"
    try:
        # Display logo
        customer_logo(0.04)

        # 1 : Collect information about needed files
        os.system("cls")
        print("*"*76 + header + "*"*76) # Display banner
        in_filename, in_filename_type, in_matrix, out_path = collect_files(args)

        # 2 : Process XLSX file
        os.system("cls")
        print("*"*76 + header + "*"*76) # Display banner
        print("[INFO] Processing XLSX file")
        convertion_matrix = process_xlsx_file(in_matrix)

        # 3 : Verifying files integrity
        # TODO : Optimize code in order to allow user to statically create models based on iface
        print("[INFO] Verifying files in directory")
        if in_filename_type == "dir":
            for filename in in_filename:
                # Check if file is a config file and what type of ISR is it
                file_valid = check_cfg_file(filename)
                if file_valid == True: # Only add valid cfg files
                    dev_type = check_dev_model(filename)
                    if dev_type != "Unknown": # Only add known device type
                        valid_files.append((filename, dev_type))
        elif in_filename_type == "file":
            # Check if file is a config file and what type of ISR is it
            file_valid = check_cfg_file(in_filename)
            if file_valid == True: # Only add valid cfg files
                dev_type = check_dev_model(in_filename)
                if dev_type != "Unknown": # Only add known device type
                    valid_files.append((in_filename, dev_type))
        print(f"==> {len(valid_files)} valid file(s) found")

        # 4 : Convert files
        print("[INFO] Converting files confg in directory")
        if len(valid_files) > 0:
            for file_params in tqdm(valid_files, desc="converting cfg file", unit="file"):
                new_file = process_file(file_params, out_path, convertion_matrix)
                new_files.append(new_file)

        # 5 : Display converted files
        display_new_files(new_files)
        input("\nPlease press Enter key to exit !")

    except SystemExit: # Occurs if help menu is displayed
        pass

    except KeyboardInterrupt: # User break
        os.system("cls")
        print("Exiting software as requested !")
        input("Please press Enter key to exit !")

    except: # Other errors
        os.system("cls")
        exception_logo()
        print("Unexpected error, software will end now !")
        print("More information about exception:", sys.exc_info()[0])
        input("Please press Enter key to exit !")